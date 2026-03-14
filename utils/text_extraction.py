from pypdf import PdfReader
from docx import Document
import openpyxl


def extract_from_pdf(file_path: str) -> str:
    """Extract text from PDF file."""
    try:
        reader = PdfReader(file_path)
        return " ".join(p.extract_text() for p in reader.pages if p.extract_text())
    except Exception as e:
        return f"Error reading PDF: {e}"


def extract_from_docx(file_path: str) -> str:
    """Extract text from Word document."""
    try:
        doc = Document(file_path)
        return "\n".join(p.text for p in doc.paragraphs if p.text)
    except Exception as e:
        return f"Error reading DOCX: {e}"


def extract_from_xlsx(file_path: str) -> str:
    """Extract text from Excel file."""
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        text_parts = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            text_parts.append(f"Sheet: {sheet}")
            for row in ws.iter_rows(values_only=True):
                row_text = " | ".join(
                    str(cell) if cell is not None else "" for cell in row
                )
                if row_text.strip():
                    text_parts.append(row_text)
        return "\n".join(text_parts)
    except Exception as e:
        return f"Error reading XLSX: {e}"


def extract_from_text(file_path: str) -> str:
    """Extract text from text file."""
    for encoding in ["utf-8", "latin-1"]:
        try:
            with open(file_path, "r", encoding=encoding) as f:
                return f.read()
        except Exception:
            continue
    return "Error reading file"


def extract_text(file_path: str) -> str:
    """Extract text based on file extension."""
    ext = file_path.lower()

    if ext.endswith(".pdf"):
        return extract_from_pdf(file_path)
    elif ext.endswith(".docx"):
        return extract_from_docx(file_path)
    elif ext.endswith(".xlsx") or ext.endswith(".xls"):
        return extract_from_xlsx(file_path)
    else:
        return extract_from_text(file_path)
